import openpyxl

def simulate_siren_arbitrage():
    wb = openpyxl.load_workbook("Arbitrage_Calculator.xlsx")
    ws = wb.active
    
    ws["B4"] = "Спот-Фьючерс"
    ws["B5"] = 1000  # Capital
    ws["B6"] = 2     # Leverage (2x is safer, or 3x)
    
    # Bybit Futures Liquidity
    ws["B7"] = 3301105.50 # Open Interest
    ws["B8"] = 30000000.00 # 24h Vol (roughly, since OI is 3.3M)
    
    # Exchange A (MEXC Spot)
    ws["B9"] = 0.1315  # Bid
    ws["B10"] = 0.1320 # Ask (Price to buy)
    
    # Exchange B (Bybit Futures)
    ws["B11"] = 0.13263 # Bid (Price to sell/short)
    ws["B12"] = 0.13267 # Ask
    
    # Funding
    ws["B13"] = 0.0     # Spot funding
    ws["B14"] = 0.186478 / 100 # Bybit funding rate (0.186478%)
    ws["B15"] = 8       # Spot interval
    ws["B16"] = 4       # Bybit interval (4h)
    
    calc_and_print(ws)

def calc_and_print(ws):
    b4 = ws["B4"].value
    b5 = ws["B5"].value
    b6 = ws["B6"].value
    b7 = ws["B7"].value
    b8 = ws["B8"].value
    b9 = ws["B9"].value
    b10 = ws["B10"].value
    b11 = ws["B11"].value
    b12 = ws["B12"].value
    b13 = ws["B13"].value
    b14 = ws["B14"].value
    b15 = ws["B15"].value
    b16 = ws["B16"].value

    spread_a = (b10 - b9) / b9
    spread_b = (b12 - b11) / b11
    price_a = b10
    price_b = b11
    pos_size = b5 / 2 * b6
    pct_oi = pos_size / b7
    pct_vol = pos_size / b8
    spread_price = (price_b - price_a) / price_a
    comm = 0.002 if b4 == "Фьючерс-Фьючерс" else 0.003
    net_spread = spread_price - comm
    
    funding_a_yield = abs(b13) if b13 < 0 else -b13
    funding_b_yield = b14 if b14 > 0 else -abs(b14)
    net_funding_daily = 24 * ((funding_a_yield / b15) + (funding_b_yield / b16))
    
    daily_income = pos_size * net_funding_daily
    payback_hours = 0 if net_spread >= 0 else abs(net_spread) / (net_funding_daily / 24)
    
    risk_liq = "ВЫСОКИЙ (крупный объем)" if (pct_oi > 0.01 or pct_vol > 0.002) else "Низкий (Безопасно)"
    risk_spread = "ПЛОХОЙ (проскальзывание)" if (spread_a > 0.002 or spread_b > 0.002) else "Отличный (Маленький)"
    
    verdict = "🟢 Входить в сделку" if (risk_liq == "Низкий (Безопасно)" and risk_spread == "Отличный (Маленький)" and payback_hours <= 12) else "🔴 Пропустить (Высокий риск)"
    
    print(f"1. Спред в стакане Биржи А: {spread_a*100:.3f}%")
    print(f"2. Спред в стакане Биржи Б: {spread_b*100:.3f}%")
    print(f"3. Цена входа А (Покупка): {price_a}")
    print(f"4. Цена входа Б (Шорт): {price_b}")
    print(f"5. Размер позиции на сторону: ${pos_size:.2f}")
    print(f"6. Доля в Открытом Интересе (OI): {pct_oi*100:.4f}%")
    print(f"7. Доля в суточном объеме: {pct_vol*100:.4f}%")
    print(f"8. Межбиржевой спред цен: {spread_price*100:+.3f}%")
    print(f"9. Чистый спред цен (с комсой): {net_spread*100:+.3f}%")
    print(f"10. Чистый фандинг в сутки (для вас): {net_funding_daily*100:+.4f}%")
    print(f"11. Доход от фандинга в день: ${daily_income:.2f}")
    print(f"12. Время окупаемости спреда (ч): {payback_hours:.1f}")
    print(f"13. Оценка риска ликвидности: {risk_liq}")
    print(f"14. Оценка спреда в стакане: {risk_spread}")
    print(f"15. Итоговый Вердикт: {verdict}")

if __name__ == "__main__":
    simulate_siren_arbitrage()
