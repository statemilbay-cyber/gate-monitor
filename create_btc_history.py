import requests
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_cycle_time(ts_ms):
    # Convert timestamp in ms to datetime in UTC
    dt = datetime.datetime.fromtimestamp(ts_ms / 1000, datetime.timezone.utc)
    # Round to the nearest hour
    dt_rounded = dt.replace(minute=0, second=0, microsecond=0)
    hour = dt_rounded.hour
    
    # Align to standard 8-hour cycles: 00:00, 08:00, 16:00 UTC
    if hour in [23, 0, 1]:
        cycle_dt = dt_rounded.replace(hour=0)
        if hour == 23:
            cycle_dt = cycle_dt + datetime.timedelta(days=1)
    elif hour in [7, 8, 9]:
        cycle_dt = dt_rounded.replace(hour=8)
    elif hour in [15, 16, 17]:
        cycle_dt = dt_rounded.replace(hour=16)
    else:
        cycle_dt = dt_rounded
        
    return cycle_dt

def fetch_bybit_history(target_date_utc):
    print("Fetching Bybit historical funding rates...")
    bybit_url = 'https://api.bybit.com/v5/market/funding/history'
    target_ts = int(target_date_utc.timestamp() * 1000)
    records = {}
    end_time = None
    
    while True:
        params = {
            'category': 'linear',
            'symbol': 'BTCUSDT',
            'limit': 200
        }
        if end_time:
            params['endTime'] = end_time
            
        try:
            r = requests.get(bybit_url, params=params, timeout=15).json()
            if r.get('retCode') != 0:
                print('Bybit API Error:', r)
                break
                
            list_data = r.get('result', {}).get('list', [])
            if not list_data:
                break
                
            for item in list_data:
                ts = int(item['fundingRateTimestamp'])
                rate = float(item['fundingRate'])
                cycle_dt = get_cycle_time(ts)
                records[cycle_dt] = rate
                
            oldest_ts = int(list_data[-1]['fundingRateTimestamp'])
            if oldest_ts <= target_ts:
                break
                
            end_time = oldest_ts - 1000
            
        except Exception as e:
            print('Error fetching from Bybit:', e)
            break
            
    return records

def fetch_binance_history(target_date_utc):
    print("Fetching Binance historical funding rates...")
    binance_url = 'https://fapi.binance.com/fapi/v1/fundingRate'
    start_ts = int(target_date_utc.timestamp() * 1000)
    records = {}
    
    # Binance limit is 1000. 1000 items is ~333 days of 8h data.
    # From Jan 1, 2026 to mid-June 2026 is about 164 days, so 492 records.
    # A single fetch is sufficient, but we can use pagination if needed.
    params = {
        'symbol': 'BTCUSDT',
        'startTime': start_ts,
        'limit': 1000
    }
    
    try:
        r = requests.get(binance_url, params=params, timeout=15).json()
        if isinstance(r, list):
            for item in r:
                ts = int(item['fundingTime'])
                rate = float(item['fundingRate'])
                cycle_dt = get_cycle_time(ts)
                records[cycle_dt] = rate
        else:
            print('Binance API Error:', r)
            
    except Exception as e:
        print('Error fetching from Binance:', e)
        
    return records

def create_excel_workbook():
    target_date = datetime.datetime(2026, 1, 1, 0, 0, 0, tzinfo=datetime.timezone.utc)
    
    # 1. Fetch data
    bybit_data = fetch_bybit_history(target_date)
    binance_data = fetch_binance_history(target_date)
    
    # Combine timestamps
    all_cycles = sorted(list(set(bybit_data.keys()).union(set(binance_data.keys()))))
    
    # Filter cycles to start from Jan 1, 2026
    start_dt_limit = datetime.datetime(2026, 1, 1, 0, 0, tzinfo=datetime.timezone.utc)
    all_cycles = [c for c in all_cycles if c >= start_dt_limit]
    
    print(f"Total cycles collected: {len(all_cycles)}")
    
    # 2. Build workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Calculator
    ws_calc = wb.active
    ws_calc.title = "Calculator"
    ws_calc.views.sheetView[0].showGridLines = True
    
    # Sheet 2: History
    ws_hist = wb.create_sheet(title="History")
    ws_hist.views.sheetView[0].showGridLines = True
    
    # Fonts
    font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True, color="666666")
    
    # Fills
    fill_blue_dark = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid") # Dark Navy
    fill_blue_light = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid") # Indigo
    fill_gray_input = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid") # Warm grey
    fill_green_output = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") # Mint green
    fill_accent = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid") # Soft blue
    
    # Borders
    thin_side = Side(style='thin', color='CFD8DC')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_bottom = Border(bottom=Side(style='medium', color='1A237E'))
    double_bottom = Border(top=thin_side, bottom=Side(style='double', color='1A237E'))
    
    # --- Populating History Sheet ---
    ws_hist.row_dimensions[1].height = 25
    headers_hist = ["Дата и время (UTC)", "Bybit Ставка фандинга", "Binance Ставка фандинга"]
    for col_idx, h in enumerate(headers_hist, 1):
        cell = ws_hist.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_blue_light
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, cycle in enumerate(all_cycles, 2):
        # Write cycle date as naive datetime for Excel compatibility
        naive_dt = cycle.replace(tzinfo=None)
        cell_date = ws_hist.cell(row=row_idx, column=1, value=naive_dt)
        cell_date.number_format = 'yyyy-mm-dd hh:mm'
        cell_date.alignment = Alignment(horizontal="center")
        cell_date.font = font_regular
        
        # Bybit Rate
        bybit_rate = bybit_data.get(cycle, 0.0)
        cell_bybit = ws_hist.cell(row=row_idx, column=2, value=bybit_rate)
        cell_bybit.number_format = '0.0000%'
        cell_bybit.alignment = Alignment(horizontal="right")
        cell_bybit.font = font_regular
        
        # Binance Rate
        binance_rate = binance_data.get(cycle, 0.0)
        cell_binance = ws_hist.cell(row=row_idx, column=3, value=binance_rate)
        cell_binance.number_format = '0.0000%'
        cell_binance.alignment = Alignment(horizontal="right")
        cell_binance.font = font_regular
        
    # Auto-adjust column widths for History
    for col in ws_hist.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_hist.column_dimensions[col_letter].width = max(max_len + 3, 20)
        
    # --- Populating Calculator Sheet ---
    ws_calc.column_dimensions['A'].width = 32
    ws_calc.column_dimensions['B'].width = 24
    ws_calc.column_dimensions['C'].width = 15
    ws_calc.column_dimensions['D'].width = 50
    
    # Title Block
    ws_calc.merge_cells("A1:D1")
    title_cell = ws_calc["A1"]
    title_cell.value = "BTC Дельта-Нейтральный Калькулятор Фандинга (2026)"
    title_cell.font = font_title
    title_cell.fill = fill_blue_dark
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_calc.row_dimensions[1].height = 40
    
    # Inputs Header
    ws_calc.merge_cells("A2:D2")
    ws_calc["A2"] = "1. ВХОДНЫЕ ПАРАМЕТРЫ (Заполнять вручную)"
    ws_calc["A2"].font = font_section
    ws_calc["A2"].fill = fill_blue_light
    ws_calc["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws_calc.row_dimensions[2].height = 24
    
    # Inputs rows
    inputs_data = [
        ("Общий капитал (USDT)", 1000, "$", "Всего средств выделено под арбитраж"),
        ("Размер позиции на фьючерсах (USDT)", "=B3/2", "$", "Выделяется под шорт-позицию (50% от капитала)"),
        ("Биржа фьючерсов", "Bybit", "текст", "Укажите биржу для шорта: Bybit или Binance"),
        ("Дата и время входа", datetime.datetime(2026, 1, 1, 8, 0, 0), "дата", "Формат: ГГГГ-ММ-ДД ЧЧ:ММ (напр. 2026-01-01 08:00)"),
        ("Вход как (Maker/Taker)", "Taker", "текст", "Спот всегда Taker (0.1%), фьючерс по выбору (Maker/Taker)"),
        ("Дата и время выхода", datetime.datetime(2026, 6, 1, 16, 0, 0), "дата", "Формат: ГГГГ-ММ-ДД ЧЧ:ММ (напр. 2026-06-01 16:00)"),
        ("Выход как (Maker/Taker)", "Maker", "текст", "Спот всегда Taker (0.1%), фьючерс по выбору (Maker/Taker)")
    ]
    
    for idx, (label, val, fmt, desc) in enumerate(inputs_data, 3):
        ws_calc.row_dimensions[idx].height = 20
        c_lbl = ws_calc.cell(row=idx, column=1, value=label)
        c_lbl.font = font_bold
        c_lbl.border = thin_border
        
        c_val = ws_calc.cell(row=idx, column=2, value=val)
        c_val.font = font_bold
        c_val.fill = fill_gray_input
        c_val.border = thin_border
        
        if fmt == "$":
            c_val.number_format = '$#,##0.00'
            c_val.alignment = Alignment(horizontal="right")
        elif fmt == "дата":
            if isinstance(val, datetime.datetime):
                c_val.value = val.replace(tzinfo=None)
            c_val.number_format = 'yyyy-mm-dd hh:mm'
            c_val.alignment = Alignment(horizontal="center")
        else:
            c_val.alignment = Alignment(horizontal="center")
            
        c_fmt = ws_calc.cell(row=idx, column=3, value=fmt)
        c_fmt.font = font_italic
        c_fmt.alignment = Alignment(horizontal="center")
        c_fmt.border = thin_border
        
        c_dsc = ws_calc.cell(row=idx, column=4, value=desc)
        c_dsc.font = font_regular
        c_dsc.border = thin_border

    # Results Header
    ws_calc.merge_cells("A10:D10")
    ws_calc["A10"] = "2. РАСЧЕТ РЕЗУЛЬТАТОВ (Вычисляется автоматически)"
    ws_calc["A10"].font = font_section
    ws_calc["A10"].fill = fill_blue_light
    ws_calc["A10"].alignment = Alignment(horizontal="left", vertical="center")
    ws_calc.row_dimensions[10].height = 24
    
    # Results rows
    results_def = [
        # Label, Formula, Format, Description
        ("Период удержания (дней)", "=B8-B6", "0.0", "Количество дней удержания позиции"),
        ("Комиссия за вход (USDT)", "=B4*0.001 + B4*IF(B7=\"Maker\", 0.0002, IF(B5=\"Bybit\", 0.00055, 0.0005))", "$", "0.1% спот taker + фьючерс (0.02% maker / 0.05-0.055% taker)"),
        ("Комиссия за выход (USDT)", "=B4*0.001 + B4*IF(B9=\"Maker\", 0.0002, IF(B5=\"Bybit\", 0.00055, 0.0005))", "$", "0.1% спот taker + фьючерс (0.02% maker / 0.05-0.055% taker)"),
        ("Всего торговых комиссий (USDT)", "=B12+B13", "$", "Сумма комиссий на вход и выход с обеих ног"),
        ("Накопленный фандинг (%)", '=IF(B5="Bybit", SUMIFS(History!B:B, History!A:A, ">="&B6, History!A:A, "<="&B8), SUMIFS(History!C:C, History!A:A, ">="&B6, History!A:A, "<="&B8))', "%", "Суммарная ставка фандинга за выбранный период"),
        ("Накопленный фандинг (USDT)", "=B4*B15", "$", "Чистый доход по фандингу в долларах"),
        ("Чистая прибыль (USDT)", "=B16-B14", "$", "Доход по фандингу минус торговые комиссии"),
        ("Чистая доходность (%)", "=B17/B3", "%", "Процент чистой прибыли к общему вложенному капиталу"),
        ("Годовая доходность (APY %)", "=B18*(365/B11)", "%", "Экстраполяция доходности в годовой процент")
    ]
    
    for offset, (label, formula, fmt, desc) in enumerate(results_def, 11):
        ws_calc.row_dimensions[offset].height = 20
        c_lbl = ws_calc.cell(row=offset, column=1, value=label)
        c_lbl.font = font_bold
        c_lbl.border = thin_border
        
        c_val = ws_calc.cell(row=offset, column=2, value=formula)
        c_val.font = font_bold
        c_val.fill = fill_green_output
        c_val.border = thin_border
        
        if fmt == "$":
            c_val.number_format = '$#,##0.00'
            c_val.alignment = Alignment(horizontal="right")
        elif fmt == "%":
            c_val.number_format = '0.0000%' if label == "Накопленный фандинг (%)" else '0.00%'
            c_val.alignment = Alignment(horizontal="right")
        elif fmt == "0.0":
            c_val.number_format = '0.0'
            c_val.alignment = Alignment(horizontal="right")
            
        c_fmt = ws_calc.cell(row=offset, column=3, value="")
        c_fmt.border = thin_border
        
        c_dsc = ws_calc.cell(row=offset, column=4, value=desc)
        c_dsc.font = font_regular
        c_dsc.border = thin_border
        
        # Highlight final performance rows
        if label in ["Чистая прибыль (USDT)", "Чистая доходность (%)", "Годовая доходность (APY %)"]:
            c_lbl.fill = fill_accent
            c_val.fill = fill_accent
            c_val.font = Font(name="Segoe UI", size=10, bold=True, color="1A237E")
            
    # Save workbook
    file_name = "BTC_Funding_History.xlsx"
    wb.save(file_name)
    print(f"Workbook saved successfully as {file_name}")

if __name__ == "__main__":
    create_excel_workbook()
