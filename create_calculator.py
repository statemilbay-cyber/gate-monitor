import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_arbitrage_calculator():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arbitrage Calculator"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True, color="555555")
    
    fill_title = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid") # Dark blue
    fill_section = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid") # Indigo
    fill_input = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid") # Cool grey
    fill_calc = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid") # Light green
    fill_decision = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid") # Soft yellow
    
    thin_border = Border(
        left=Side(style='thin', color='CFD8DC'),
        right=Side(style='thin', color='CFD8DC'),
        top=Side(style='thin', color='CFD8DC'),
        bottom=Side(style='thin', color='CFD8DC')
    )
    
    # Title Block
    ws.merge_cells("A1:C1")
    ws["A1"] = "Калькулятор арбитража ставок финансирования"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Section 1: Inputs
    ws.merge_cells("A2:C2")
    ws["A2"] = "1. ВХОДНЫЕ ДАННЫЕ (Заполнять вручную)"
    ws["A2"].font = font_header
    ws["A2"].fill = fill_section
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24
    
    inputs = [
        ("Монета", "HOME", "Тикер монеты (напр. HOME, BEAT)"),
        ("Тип арбитража", "Спот-Фьючерс", "Фьючерс-Фьючерс или Спот-Фьючерс"),
        ("Ваш общий капитал (USDT)", 1000, "Сумма, выделенная на обе стороны вместе"),
        ("Плечо (Leverage)", 3, "Рабочее кредитное плечо на фьючерсах (рекомендуется 2-3)"),
        ("Открытый интерес в USD (OI)", 8970000, "Сумма всех открытых позиций на бирже фьючерсов"),
        ("Суточный объем торгов (24h Vol)", 11020000, "24-часовой объем торгов на фьючерсной бирже в USD"),
        ("Биржа А - Лучший BUY (Bid) цена", 0.02984, "Лучшая цена покупки лимитным ордером (покупатели) на Бирже А"),
        ("Биржа А - Лучший SELL (Ask) цена", 0.02986, "Лучшая цена продажи лимитным ордером (продавцы) на Бирже А"),
        ("Биржа Б - Лучший BUY (Bid) цена", 0.02985, "Лучшая цена покупки лимитным ордером (покупатели) на Бирже Б"),
        ("Биржа Б - Лучший SELL (Ask) цена", 0.02987, "Лучшая цена продажи лимитным ордером (продавцы) на Бирже Б"),
        ("Ставка фандинга Биржи А / Спота (%)", 0.0, "Ставка фандинга в % (для спота = 0)"),
        ("Ставка фандинга Биржи Б / Фьюч (%)", -0.006031, "Ставка фандинга в % на второй бирже (e.g. -0.6031% = -0.006031)"),
        ("Интервал фандинга Биржи А (часы)", 8, "Интервал списания фандинга на первой бирже (для спота = 8)"),
        ("Интервал фандинга Биржи Б (часы)", 8, "Интервал списания фандинга на фьючерсах (обычно 8, 4 или 1)")
    ]
    
    row_idx = 3
    for label, val, desc in inputs:
        ws.cell(row=row_idx, column=1, value=label).font = font_bold
        ws.cell(row=row_idx, column=1).border = thin_border
        
        cell_val = ws.cell(row=row_idx, column=2, value=val)
        cell_val.font = font_regular
        cell_val.fill = fill_input
        cell_val.border = thin_border
        cell_val.alignment = Alignment(horizontal="right")
        
        # Formats
        if "общий капитал" in label or "интерес" in label or "объем" in label:
            cell_val.number_format = '$#,##0.00'
        elif "цена" in label or "Цена" in label:
            cell_val.number_format = '0.000000'
        elif "(%)" in label:
            cell_val.number_format = '0.00%'
        
        ws.cell(row=row_idx, column=3, value=desc).font = font_italic
        ws.cell(row=row_idx, column=3).border = thin_border
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    # Spacer row
    ws.row_dimensions[row_idx].height = 10
    row_idx += 1
    
    # Section 2: Calculations
    ws.merge_cells(f"A{row_idx}:C{row_idx}")
    ws[f"A{row_idx}"] = "2. АВТОМАТИЧЕСКИЕ РАСЧЕТЫ"
    ws[f"A{row_idx}"].font = font_header
    ws[f"A{row_idx}"].fill = fill_section
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_idx].height = 24
    row_idx += 1
    
    calcs = [
        ("Спред в стакане Биржи А (%)", "=(B10-B9)/B9", "Спред в стакане Биржи А: (Ask - Bid) / Bid"),
        ("Спред в стакане Биржи Б (%)", "=(B12-B11)/B11", "Спред в стакане Биржи Б: (Ask - Bid) / Bid"),
        ("Цена входа на Бирже А (Покупка)", "=B10", "Вы покупаете по цене лучшей продажи (Ask)"),
        ("Цена входа на Бирже Б (Шорт/Продажа)", "=B11", "Вы продаете по цене лучшей покупки (Bid)"),
        ("Размер позиции на одну сторону (USD)", "=B5/2*B6", "Капитал на одну сторону с учетом плеча (Капитал/2 * Плечо)"),
        ("Доля в Открытом интересе (OI) (%)", "=B23/B7", "Рекомендуется < 0.5% (более 1% опасно для входа/выхода)"),
        ("Доля в суточном объеме (%)", "=B23/B8", "Рекомендуется < 0.1% (более 0.2% приведет к проскальзыванию)"),
        ("Межбиржевой спред цен (%)", "=(B22-B21)/B21", "Разница цен между биржами (положительный спред — хорошо)"),
        ("Ориентировочная комиссия на круг (%)", '=IF(B4="Фьючерс-Фьючерс", 0.002, 0.003)', "Комиссия на открытие + закрытие позиций (Taker-fee)"),
        ("Чистый спред с учетом комиссий (%)", "=B26-B27", "Спред за вычетом торговых комиссий обеих бирж"),
        ("Чистый фандинг в сутки (%)", "=24*((IF(B13<0, ABS(B13), -B13)/B15) + (IF(B14>0, B14, -ABS(B14))/B16))", "Суточная чистая прибыль от фандинга (с учетом знака ставки)"),
        ("Доход от фандинга в день (USD)", "=B23*B29", "Прогнозируемая прибыль в день в долларах на ваш рабочий объем"),
        ("Время окупаемости спреда (часы)", "=IF(B28>=0, 0, ABS(B28)/(B29/24))", "За сколько часов фандинг отобьет минусовой спред на входе"),
        ("Оценка риска ликвидности", '=IF(OR(B24>0.01, B25>0.002), "ВЫСОКИЙ (крупный объем)", "Низкий (Безопасно)")', "Проверка объема позиции к ликвидности стакана"),
        ("Оценка спреда в стакане", '=IF(OR(B19>0.002, B20>0.002), "ПЛОХОЙ (проскальзывание)", "Отличный (Маленький)")' , "Проверка плотности ордеров в стаканах обеих бирж")
    ]
    
    for label, formula, desc in calcs:
        ws.cell(row=row_idx, column=1, value=label).font = font_bold
        ws.cell(row=row_idx, column=1).border = thin_border
        
        cell_formula = ws.cell(row=row_idx, column=2, value=formula)
        cell_formula.font = font_bold
        cell_formula.fill = fill_calc
        cell_formula.border = thin_border
        cell_formula.alignment = Alignment(horizontal="right")
        
        # Formats
        if "объем" in label or "день" in label or "Размер" in label:
            cell_formula.number_format = '$#,##0.00'
        elif "Цена" in label or "цена" in label:
            cell_formula.number_format = '0.000000'
        elif "(%)" in label:
            cell_formula.number_format = '0.00%'
        elif "часы" in label:
            cell_formula.number_format = '0.0'
            
        ws.cell(row=row_idx, column=3, value=desc).font = font_italic
        ws.cell(row=row_idx, column=3).border = thin_border
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    # Spacer row
    ws.row_dimensions[row_idx].height = 10
    row_idx += 1
    
    # Section 3: Decision
    ws.merge_cells(f"A{row_idx}:C{row_idx}")
    ws[f"A{row_idx}"] = "3. АВТОМАТИЧЕСКИЙ ВЕРДИКТ"
    ws[f"A{row_idx}"].font = font_header
    ws[f"A{row_idx}"].fill = fill_section
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_idx].height = 24
    row_idx += 1
    
    # The Final Decision
    ws.cell(row=row_idx, column=1, value="Решение по сделке:").font = Font(name="Segoe UI", size=11, bold=True)
    ws.cell(row=row_idx, column=1).border = thin_border
    
    decision_formula = f'=IF(AND(B{row_idx-4}="Низкий (Безопасно)", B{row_idx-3}="Отличный (Маленький)", B{row_idx-5}<=12), "🟢 Входить в сделку", "🔴 Пропустить (Высокий риск или долгая окупаемость)")'
    cell_decision = ws.cell(row=row_idx, column=2, value=decision_formula)
    cell_decision.font = Font(name="Segoe UI", size=11, bold=True)
    cell_decision.fill = fill_decision
    cell_decision.border = thin_border
    cell_decision.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.cell(row=row_idx, column=3, value="Сделка одобряется, если ликвидность безопасна, спред в стакане узкий и окупаемость входа < 12 часов").font = font_italic
    ws.cell(row=row_idx, column=3).border = thin_border
    ws.row_dimensions[row_idx].height = 30
    
    # Set Column Widths
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 75
    
    wb.save("Arbitrage_Calculator.xlsx")
    print("Excel Arbitrage Calculator successfully generated!")

if __name__ == "__main__":
    create_arbitrage_calculator()

