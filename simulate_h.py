import openpyxl

def simulate_h_arbitrage():
    # Load workbook
    wb = openpyxl.load_workbook("Arbitrage_Calculator.xlsx")
    ws = wb.active
    
    # Capital and Leverage
    ws["B5"] = 1000
    ws["B6"] = 3
    
    # Bybit Futures Liquidity
    ws["B7"] = 21101233.55 # Open Interest
    ws["B8"] = 80814844.46 # 24h Vol
    
    print("=== SIMULATION 1: Spot-Futures (MEXC Spot -> Bybit Futures) ===")
    ws["B4"] = "Спот-Фьючерс"
    # Exchange A (MEXC Spot)
    ws["B9"] = 0.22514  # Bid
    ws["B10"] = 0.22563 # Ask (Price to buy)
    # Exchange B (Bybit Futures)
    ws["B11"] = 0.23954 # Bid (Price to sell/short)
    ws["B12"] = 0.23965 # Ask
    # Funding
    ws["B13"] = 0.0     # Spot funding (0)
    ws["B14"] = 0.475287 / 100 # Bybit funding (0.475287%)
    ws["B15"] = 8       # Spot funding interval (8)
    ws["B16"] = 1       # Bybit funding interval (1h)
    
    # Save temporarily to trigger formulas or calculate manually in Python
    # Since openpyxl doesn't evaluate formulas on save, we will calculate them in Python using the exact same logic:
    calc_and_print(ws)

    print("\n=== SIMULATION 2: Futures-Futures (MEXC Futures -> Bybit Futures) ===")
    ws["B4"] = "Фьючерс-Фьючерс"
    # Exchange A (MEXC Futures)
    ws["B9"] = 0.23407  # Bid
    ws["B10"] = 0.23417 # Ask (Price to buy)
    # Exchange B (Bybit Futures)
    ws["B11"] = 0.23954 # Bid (Price to sell/short)
    ws["B12"] = 0.23965 # Ask
    # Funding
    ws["B13"] = 0.0012 / 100   # MEXC Futures funding (0.0012%)
    ws["B14"] = 0.475287 / 100  # Bybit Futures funding (0.475287%)
    ws["B15"] = 8       # MEXC interval
    ws["B16"] = 1       # Bybit interval
    
    calc_and_print(ws)

def calc_and_print(ws):
    # Retrieve inputs
    b4 = ws["B4"].value
    b5 = ws["B5"].value
    b6 = ws["B6"].value
    b7 = ws["B7"].value
    b8 = ws["B8"].value
    b9 = ws["B9"].value
    b10 = ws["B10"].value
    b11 = ws["B11"].value
    b12 = ws["B12"].value
    b13 = ws["B13"].value
    b14 = ws["B14"].value
    b15 = ws["B15"].value
    b16 = ws["B16"].value

    # Formulas equivalent in Python
    spread_a = (b10 - b9) / b9
    spread_b = (b12 - b11) / b11
    price_a = b10
    price_b = b11
    pos_size = b5 / 2 * b6
    pct_oi = pos_size / b7
    pct_vol = pos_size / b8
    spread_price = (price_b - price_a) / price_a
    comm = 0.002 if b4 == "Фьючерс-Фьючерс" else 0.003
    net_spread = spread_price - comm
    
    # Funding logic
    funding_a_yield = abs(b13) if b13 < 0 else -b13
    funding_b_yield = b14 if b14 > 0 else -abs(b14)
    net_funding_daily = 24 * ((funding_a_yield / b15) + (funding_b_yield / b16))
    
    daily_income = pos_size * net_funding_daily
    payback_hours = 0 if net_spread >= 0 else abs(net_spread) / (net_funding_daily / 24)
    
    risk_liq = "ВЫСОКИЙ (крупный объем)" if (pct_oi > 0.01 or pct_vol > 0.002) else "Низкий (Безопасно)"
    risk_spread = "ПЛОХОЙ (проскальзывание)" if (spread_a > 0.002 or spread_b > 0.002) else "Отличный (Маленький)"
    
    verdict = "🟢 Входить в сделку" if (risk_liq == "Низкий (Безопасно)" and risk_spread == "Отличный (Маленький)" and payback_hours <= 12) else "🔴 Пропустить (Высокий риск)"
    
    print(f"1. Спред в стакане Биржи А: {spread_a*100:.3f}%")
    print(f"2. Спред в стакане Биржи Б: {spread_b*100:.3f}%")
    print(f"3. Цена входа А: {price_a}")
    print(f"4. Цена входа Б: {price_b}")
    print(f"5. Размер позиции на сторону: ${pos_size:.2f}")
    print(f"6. Доля в Открытом Интересе (OI): {pct_oi*100:.4f}%")
    print(f"7. Доля в суточном объеме: {pct_vol*100:.4f}%")
    print(f"8. Межбиржевой спред цен: {spread_price*100:+.3f}%")
    print(f"9. Комиссии на круг: {comm*100:.2f}%")
    print(f"10. Чистый спред цен (с комсой): {net_spread*100:+.3f}%")
    print(f"11. Чистый фандинг в сутки (для вас): {net_funding_daily*100:+.4f}%")
    print(f"12. Доход от фандинга в день: ${daily_income:.2f}")
    print(f"13. Время окупаемости спреда (ч): {payback_hours:.1f}")
    print(f"14. Оценка риска ликвидности: {risk_liq}")
    print(f"15. Оценка спреда в стакане: {risk_spread}")
    print(f"16. Итоговый Вердикт: {verdict}")

if __name__ == "__main__":
    simulate_h_arbitrage()
